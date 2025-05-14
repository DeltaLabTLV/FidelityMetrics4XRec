#!/usr/bin/env python
# coding: utf-8

import pandas as pd
import numpy as np
import os
# os.environ['KMP_DUPLICATE_LIB_OK'] = 'True' # Keep commented
# export_dir = os.getcwd() # Removed top-level, unused one
from pathlib import Path
import pickle
from collections import defaultdict
import torch
# import torch.nn as nn # Removed, not directly used after LXR Explainer local class removal
# import matplotlib.pyplot as plt # Keep commented
# import ipynb # Remove
from tqdm import tqdm
# from concurrent.futures import ProcessPoolExecutor # Removed, unused

data_name = "ML1M" ### Can be ML1M, Yahoo, Pinterest
recommender_name = "MLP" ### Can be MLP, VAE, NCF

DP_DIR = Path("processed_data", data_name) 
export_dir = Path(os.getcwd()) # Ensuring this local export_dir is active for checkpoints_path
files_path = Path("/storage/mikhail/PI4Rec", DP_DIR)
checkpoints_path = Path(export_dir.parent, "checkpoints") # Ensuring this uses the local export_dir.parent
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

output_type_dict = {
    "VAE":"multiple",
    "MLP":"single",
    "NCF": "single",
    ("Pinterest","MLP"):Path(checkpoints_path, "MLP_Pinterest_0.0062_512_21_0.pt"),
    ("Pinterest","NCF"):Path(checkpoints_path, "NCF2_Pinterest_9e-05_32_9_10.pt")
}

num_users_dict = {
    "ML1M":6037,
    "Yahoo":13797, 
    "Pinterest":19155}

num_items_dict = {
    "ML1M":3381,
    "Yahoo":4604, 
    "Pinterest":9362}

recommender_path_dict = {
    ("ML1M","VAE"): Path(checkpoints_path, "VAE_ML1M_0.0007_128_10.pt"),
    ("ML1M","MLP"):Path(checkpoints_path, "MLP1_ML1M_0.0076_256_7.pt"),
    ("ML1M","NCF"):Path(checkpoints_path, "NCF_ML1M_5e-05_64_16.pt"),
    
    ("Yahoo","VAE"): Path(checkpoints_path, "VAE_Yahoo_0.0001_128_13.pt"),
    ("Yahoo","MLP"):Path(checkpoints_path, "MLP2_Yahoo_0.0083_128_1.pt"),
    ("Yahoo","NCF"):Path(checkpoints_path, "NCF_Yahoo_0.001_64_21_0.pt"),
    
    ("Pinterest","VAE"): Path(checkpoints_path, "VAE_Pinterest_12_18_0.0001_256.pt"),
    ("Pinterest","MLP"):Path(checkpoints_path, "MLP_Pinterest_0.0062_512_21_0.pt"),
    ("Pinterest","NCF"):Path(checkpoints_path, "NCF2_Pinterest_9e-05_32_9_10.pt"),
}

hidden_dim_dict = {
    ("ML1M","VAE"): None,
    ("ML1M","MLP"): 32,
    ("ML1M","NCF"): 8,

    ("Yahoo","VAE"): None,
    ("Yahoo","MLP"):32,
    ("Yahoo","NCF"):8,
    
    ("Pinterest","VAE"): None,
    ("Pinterest","MLP"):512,
    ("Pinterest","NCF"): 64,
}

output_type = output_type_dict[recommender_name] ### Can be single, multiple
num_users = num_users_dict[data_name] 
num_items = num_items_dict[data_name] 

hidden_dim = hidden_dim_dict[(data_name,recommender_name)]
recommender_path = recommender_path_dict[(data_name,recommender_name)]

train_data = pd.read_csv(Path(files_path,f'train_data_{data_name}.csv'), index_col=0)
test_data = pd.read_csv(Path(files_path,f'test_data_{data_name}.csv'), index_col=0)
static_test_data = pd.read_csv(Path(files_path,f'static_test_data_{data_name}.csv'), index_col=0)
with open(Path(files_path,f'pop_dict_{data_name}.pkl'), 'rb') as f:
    pop_dict = pickle.load(f)
train_array = train_data.to_numpy()
test_array = test_data.to_numpy()
items_array = np.eye(num_items)
all_items_tensor = torch.Tensor(items_array).to(device)

test_array = static_test_data.iloc[:,:-2].to_numpy()
with open(Path(files_path, f'jaccard_based_sim_{data_name}.pkl'), 'rb') as f:
    jaccard_dict = pickle.load(f) 
with open(Path(files_path, f'cosine_based_sim_{data_name}.pkl'), 'rb') as f:
    cosine_dict = pickle.load(f) 
with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
    pop_dict = pickle.load(f) 
with open(Path(files_path, f'item_to_cluster_{recommender_name}_{data_name}.pkl'), 'rb') as f:
    item_to_cluster = pickle.load(f) 
with open(Path(files_path, f'shap_values_{recommender_name}_{data_name}.pkl'), 'rb') as f:
    shap_values= pickle.load(f) 
for i in range(num_items):
    for j in range(i, num_items):
        jaccard_dict[(j,i)]= jaccard_dict[(i,j)]
        cosine_dict[(j,i)]= cosine_dict[(i,j)]
        pop_array = np.zeros(len(pop_dict))
for key, value in pop_dict.items():
    pop_array[key] = value
kw_dict = {
    'device': device,
    'num_items': num_items,
    'num_features': num_items,
    'pop_array': pop_array,
    'all_items_tensor': all_items_tensor,
    'static_test_data': static_test_data,
    'items_array': items_array,
    'output_type': output_type,
    'recommender_name': recommender_name
}

# import os # Already imported at the top, remove redundant import

#os.chdir('/storage/mikhail/PI4Rec/code') # Remove
# print(os.getcwd()) # Remove

# sys.path.append('../baselines') # Remove this line

from .lime import explain_instance_lime, explain_instance_lire, distance_to_proximity # NEW

# These are from the local ./code directory
from .help_functions import (
    load_recommender, get_user_recommended_item, recommender_run,
    get_top_k, get_index_in_the_list, get_ndcg,
    load_lxr_explainer, find_lxr_mask
)
# importlib.reload(ipynb.fs.defs.help_functions) # Remove

# from .recommenders_architecture import MLP, VAE, NCF, MLP_model, GMF_model # Keep commented
# importlib.reload(ipynb.fs.defs.recommenders_architecture) # Remove

# VAE_config= { # This was removed in a previous step, ensure it stays that way

def find_pop_mask(x, item_id):
    user_hist = torch.Tensor(x).to(device) # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    item_pop_dict = {}
    
    for i,j in enumerate(user_hist>0):
        if j:
            item_pop_dict[i]=pop_array[i] # add the pop of the item to the dictionary
            
    return item_pop_dict

#User based similarities using Jaccard
def find_jaccard_mask(x, item_id, user_based_Jaccard_sim):
    user_hist = x # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    item_jaccard_dict = {}
    for i,j in enumerate(user_hist>0):
        if j:
            if (i,item_id) in user_based_Jaccard_sim:
                item_jaccard_dict[i]=user_based_Jaccard_sim[(i,item_id)] # add Jaccard similarity between items
            else:
                item_jaccard_dict[i] = 0            

    return item_jaccard_dict

#Cosine based similarities between users and items
def find_cosine_mask(x, item_id, item_cosine):
    user_hist = x # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    item_cosine_dict = {}
    for i,j in enumerate(user_hist>0):
        if j:
            if (i,item_id) in item_cosine:
                item_cosine_dict[i]=item_cosine[(i,item_id)]
            else:
                item_cosine_dict[i]=0

    return item_cosine_dict

def find_fia_mask(user_tensor, item_tensor, item_id, recommender):
    y_pred = recommender_run(user_tensor, recommender, item_tensor, item_id, **kw_dict).to(device)
    items_fia = {}
    user_hist = user_tensor.cpu().detach().numpy().astype(int)
    
    for i in range(num_items):
        if(user_hist[i] == 1):
            user_hist[i] = 0
            user_tensor = torch.FloatTensor(user_hist).to(device)
            y_pred_without_item = recommender_run(user_tensor, recommender, item_tensor, item_id, 'single', **kw_dict).to(device)
            infl_score = y_pred - y_pred_without_item
            items_fia[i] = infl_score
            user_hist[i] = 1

    return items_fia

def find_shapley_mask(user_tensor, user_id, model, shap_values, item_to_cluster):
    item_shap = {}
    shapley_values = shap_values[shap_values[:, 0].astype(int) == user_id][:,1:]
    user_vector = user_tensor.cpu().detach().numpy().astype(int)

    for i in np.where(user_vector.astype(int) == 1)[0]:
        items_cluster = item_to_cluster[i]
        item_shap[i] = shapley_values.T[int(items_cluster)][0]

    return item_shap  

def find_accent_mask(user_tensor, user_id, item_tensor, item_id, recommender_model, top_k):
   
    items_accent = defaultdict(float)
    factor = top_k - 1
    user_accent_hist = user_tensor.cpu().detach().numpy().astype(int)

    #Get topk items
    sorted_indices = list(get_top_k(user_tensor, user_tensor, recommender_model, **kw_dict).keys())
    
    if top_k == 1:
        # When k=1, return the index of the first maximum value
        top_k_indices = [sorted_indices[0]]
    else:
        top_k_indices = sorted_indices[:top_k]
   

    for iteration, item_k_id in enumerate(top_k_indices):

        # Set topk items to 0 in the user's history
        user_accent_hist[item_k_id] = 0
        user_tensor = torch.FloatTensor(user_accent_hist).to(device)
       
        item_vector = items_array[item_k_id]
        item_tensor = torch.FloatTensor(item_vector).to(device)
              
        # Check influence of the items in the history on this specific item in topk
        fia_dict = find_fia_mask(user_tensor, item_tensor, item_k_id, recommender_model)
         
        # Sum up all differences between influence on top1 and other topk values
        if not iteration:
            for key in fia_dict.keys():
                items_accent[key] *= factor
        else:
            for key in fia_dict.keys():
                items_accent[key] -= fia_dict[key]
       
    for key in items_accent.keys():
        items_accent[key] *= -1    

    return items_accent

def single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items_global, recommender_model_local, user_id = None, mask_type = None):
    user_hist_size = np.sum(user_vector)

    if mask_type == 'lime':
        # Consistent with create_dictionaries.py, maps to POS LIME for this function's structure
        POS_sim_items = explain_instance_lime(
            user_vector_np=user_vector,
            item_id_to_explain=item_id,
            recommender_model=recommender_model_local,
            all_items_tensor_global=all_items_tensor, # Global from script
            kw_dict_global=kw_dict, # Global from script
            min_pert=50, max_pert=100, num_perturbations_lime=150,
            kernel_fn=distance_to_proximity,
            num_features_to_select=user_hist_size, 
            feature_selection_method='highest_weights',
            explanation_method='POS',
            random_state_lime=item_id
        )
    elif mask_type == 'lire':
        POS_sim_items = explain_instance_lire(
            user_vector_np=user_vector,
            item_id_to_explain=item_id,
            recommender_model=recommender_model_local,
            all_items_tensor_global=all_items_tensor, # Global from script
            train_array_global=train_array, # Global from script
            kw_dict_global=kw_dict, # Global from script
            num_perturbations_lire=200,
            proba_lire=0.1,
            kernel_fn=distance_to_proximity,
            num_features_to_select=user_hist_size, 
            feature_selection_method='highest_weights',
            explanation_method='POS',
            random_state_lime=item_id
        )
    else:
        # ... (other mask types: pop, jaccard, cosine, shap, accent, lxr - these remain unchanged)
        if mask_type == 'pop':
            sim_items = find_pop_mask(user_vector, item_id) # Pass user_vector (numpy) to find_pop_mask if it expects numpy
        elif mask_type == 'jaccard':
            sim_items = find_jaccard_mask(user_vector, item_id, jaccard_dict)
        elif mask_type == 'cosine':
            sim_items = find_cosine_mask(user_vector, item_id, cosine_dict)
        elif mask_type == 'shap':
            sim_items = find_shapley_mask(user_tensor, user_id, recommender_model_local, shap_values, item_to_cluster)
        elif mask_type == 'accent':
            sim_items = find_accent_mask(user_tensor, user_id, item_tensor, item_id, recommender_model_local, 5)
        elif mask_type == 'lxr':
            lxr_explainer_instance = load_lxr_explainer(
                data_name, recommender_name, checkpoints_path, num_items, device
            )
            sim_items = find_lxr_mask(user_tensor, item_tensor, lxr_explainer_instance) # Call imported find_lxr_mask

        POS_sim_items = list(sorted(sim_items.items(), key=lambda item: item[1],reverse=True))[:user_hist_size]
        
    return POS_sim_items

class MetricsBaselines:
    def __init__(self, data_name, recommender_name):
        self.data_name = data_name
        self.recommender_name = recommender_name
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.setup_data_and_recommender()

    def setup_data_and_recommender(self):
        # Set up all necessary data and variables
        DP_DIR = Path("processed_data", self.data_name)
        self.files_path = Path(export_dir.parent, DP_DIR)
        self.num_users = num_users_dict[self.data_name]
        self.num_items = num_items_dict[self.data_name]
        
        self.test_data = pd.read_csv(Path(self.files_path, f'test_data_{self.data_name}.csv'), index_col=0)
        self.test_array = self.test_data.to_numpy()
        self.items_array = np.eye(self.num_items)
        
        with open(Path(self.files_path, f'pop_dict_{self.data_name}.pkl'), 'rb') as f:
            self.pop_dict = pickle.load(f)
        
        # Load other necessary data (jaccard_dict, cosine_dict, item_to_cluster, shap_values)
        
        self.kw_dict = {
            'device': self.device,
            'num_items': self.num_items,
            'num_features': self.num_items,
            'demographic': False,
            'pop_array': np.array([self.pop_dict.get(i, 0) for i in range(self.num_items)]),
            'all_items_tensor': torch.eye(self.num_items).to(self.device),
            'static_test_data': self.test_data,
            'items_array': self.items_array,
            'output_type': output_type_dict[self.recommender_name],
            'recommender_name': self.recommender_name,
            'files_path': self.files_path
        }
        
        self.recommender = self.load_recommender()

def process_user(user_index, test_array, test_data, recommender, kw_dict):
    try:
        user_vector = test_array[user_index]
        user_tensor = torch.FloatTensor(user_vector).to(kw_dict['device'])
        user_id = int(test_data.index[user_index])

        item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
        item_vector = kw_dict['items_array'][item_id]
        item_tensor = torch.FloatTensor(item_vector).to(kw_dict['device'])

        user_vector[item_id] = 0
        user_tensor[item_id] = 0

        results = {}
        for method in ['pop', 'jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']:
            results[method] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, kw_dict['num_items'], recommender, mask_type=method, user_id=user_id if method == 'shap' else None)

        return user_id, results
    except Exception as e:
        print(f"Error processing user {user_id}: {str(e)}")
        return None

def single_user_metrics(user_vector, user_tensor, item_id, item_tensor, num_of_bins, recommender_model, expl_dict, **kw_dict):
    '''
    This function takes the explanation dictionary as input.
    It iteratively removes items from the user's history based on their explanation scores
    and calculates metrics for the resulting counterfactual user vector.
    '''
    POS_masked = user_tensor
    NEG_masked = user_tensor
    POS_masked[item_id] = 0
    NEG_masked[item_id] = 0
    user_hist_size = np.sum(user_vector)
    
    bins = [0] + [len(x) for x in np.array_split(np.arange(user_hist_size), num_of_bins, axis=0)]
    
    # Initialize arrays for both POS and NEG metrics
    POS_at_1 = [0] * (len(bins))
    POS_at_5 = [0] * (len(bins))
    POS_at_10 = [0] * (len(bins))
    POS_at_20 = [0] * (len(bins))
    
    NEG_at_1 = [0] * (len(bins))
    NEG_at_5 = [0] * (len(bins))
    NEG_at_10 = [0] * (len(bins))
    NEG_at_20 = [0] * (len(bins))
    
    DEL = [0] * (len(bins))
    INS = [0] * (len(bins))
    NDCG = [0] * (len(bins))
    
    POS_sim_items = expl_dict
    NEG_sim_items = list(sorted(dict(POS_sim_items).items(), key=lambda item: item[1], reverse=False))
    
    total_items = 0
    for i in range(len(bins)):
        total_items += bins[i]
        
        # Process POS masks
        POS_masked = torch.zeros_like(user_tensor, dtype=torch.float32, device=device)
        for j in POS_sim_items[:total_items]:
            POS_masked[j[0]] = 1
        POS_masked = user_tensor - POS_masked  # remove the masked items
        
        # Process NEG masks
        NEG_masked = torch.zeros_like(user_tensor, dtype=torch.float32, device=device)
        for j in NEG_sim_items[:total_items]:
            NEG_masked[j[0]] = 1
        NEG_masked = user_tensor - NEG_masked  # remove the masked items
        
        # Get rankings for both POS and NEG
        POS_ranked_list = get_top_k(POS_masked, user_tensor, recommender_model, **kw_dict)
        
        if item_id in list(POS_ranked_list.keys()):
            POS_index = list(POS_ranked_list.keys()).index(item_id) + 1
        else:
            POS_index = num_items
        NEG_index = get_index_in_the_list(NEG_masked, user_tensor, item_id, recommender_model, **kw_dict) + 1
        
        # Calculate POS metrics
        POS_at_1[i] = 1 if POS_index <= 1 else 0
        POS_at_5[i] = 1 if POS_index <= 5 else 0
        POS_at_10[i] = 1 if POS_index <= 10 else 0
        POS_at_20[i] = 1 if POS_index <= 20 else 0
        
        # Calculate NEG metrics
        NEG_at_1[i] = 1 if NEG_index <= 1 else 0
        NEG_at_5[i] = 1 if NEG_index <= 5 else 0
        NEG_at_10[i] = 1 if NEG_index <= 10 else 0
        NEG_at_20[i] = 1 if NEG_index <= 20 else 0
        
        # Calculate other metrics
        DEL[i] = float(recommender_run(POS_masked, recommender_model, item_tensor, item_id, **kw_dict).detach().cpu().numpy())
        INS[i] = float(recommender_run(user_tensor-POS_masked, recommender_model, item_tensor, item_id, **kw_dict).detach().cpu().numpy())
        NDCG[i] = get_ndcg(list(POS_ranked_list.keys()), item_id, **kw_dict)
    
    res = [DEL, INS, NDCG, 
           POS_at_5, POS_at_10, POS_at_20,
           NEG_at_5, NEG_at_10, NEG_at_20]
    
    for i in range(len(res)):
        res[i] = np.array(res[i])
    
    return res

def eval_one_expl_type(expl_name):
    print(f' ============ Start explaining {data_name} {recommender_name} by {expl_name} ============')
    
    # Load the appropriate explanation dictionary
    if expl_name == 'PI_base':
        with open(Path(files_path, f'{recommender_name}_PI_base_expl_dict.pkl'), 'rb') as handle:
            expl_dict = pickle.load(handle)
    else:
        with open(Path(files_path,f'{recommender_name}_{expl_name}_expl_dict.pkl'), 'rb') as handle:
            expl_dict = pickle.load(handle)
    
    recommender.eval()
    
    num_of_bins = 10  # Fixed number of bins for all users
    
    # Initialize arrays for all metrics
    metrics = {
        'DEL': np.zeros(num_of_bins + 1),
        'INS': np.zeros(num_of_bins + 1),
        'NDCG': np.zeros(num_of_bins + 1),
        'POS_at_5': np.zeros(num_of_bins + 1),
        'POS_at_10': np.zeros(num_of_bins + 1),
        'POS_at_20': np.zeros(num_of_bins + 1),
        'NEG_at_5': np.zeros(num_of_bins + 1),
        'NEG_at_10': np.zeros(num_of_bins + 1),
        'NEG_at_20': np.zeros(num_of_bins + 1)
    }

    with torch.no_grad():
        for i in tqdm(range(test_array.shape[0])):
            user_vector = test_array[i]
            user_tensor = torch.FloatTensor(user_vector).to(device)
            user_id = int(test_data.index[i])

            item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
            item_vector = items_array[item_id]
            item_tensor = torch.FloatTensor(item_vector).to(device)

            user_vector[item_id] = 0
            user_tensor[item_id] = 0

            user_expl = expl_dict[user_id]

            res = single_user_metrics(user_vector, user_tensor, item_id, item_tensor, num_of_bins, recommender, user_expl, **kw_dict)
            
            # Ensure all arrays have the same length before adding
            for j in range(len(res)):
                if len(res[j]) != len(metrics['DEL']):
                    res[j] = np.interp(np.linspace(0, 1, len(metrics['DEL'])), 
                                     np.linspace(0, 1, len(res[j])), 
                                     res[j])
            
            # Map results to metrics dictionary
            metrics['DEL'] += res[0]
            metrics['INS'] += res[1]
            metrics['NDCG'] += res[2]
            metrics['POS_at_5'] += res[3]
            metrics['POS_at_10'] += res[4]
            metrics['POS_at_20'] += res[5]
            metrics['NEG_at_5'] += res[6]
            metrics['NEG_at_10'] += res[7]
            metrics['NEG_at_20'] += res[8]

    a = test_array.shape[0]

    # Print all metrics
    for metric_name, values in metrics.items():
        print(f'{metric_name}_{expl_name}: ', np.mean(values)/a)

    # Return normalized metrics
    return {metric_name: values/a for metric_name, values in metrics.items()}

def run_all_baselines(data_name, recommender_name):
    global num_users, num_items, device, kw_dict, recommender, test_array, test_data, items_array, jaccard_dict, cosine_dict, pop_dict, item_to_cluster, shap_values

    # Update global variables for the current dataset and recommender
    num_users = num_users_dict[data_name]
    num_items = num_items_dict[data_name]
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # Load dataset-specific files
    DP_DIR = Path("processed_data", data_name)
    files_path = Path(export_dir.parent, DP_DIR)
    test_data = pd.read_csv(Path(files_path, f'test_data_{data_name}.csv'), index_col=0)
    test_array = test_data.to_numpy()
    items_array = np.eye(num_items)

    with open(Path(files_path, f'jaccard_based_sim_{data_name}.pkl'), 'rb') as f:
        jaccard_dict = pickle.load(f)
    with open(Path(files_path, f'cosine_based_sim_{data_name}.pkl'), 'rb') as f:
        cosine_dict = pickle.load(f)
    with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
        pop_dict = pickle.load(f)
    with open(Path(files_path, f'item_to_cluster_{recommender_name}_{data_name}.pkl'), 'rb') as f:
        item_to_cluster = pickle.load(f)
    with open(Path(files_path, f'shap_values_{recommender_name}_{data_name}.pkl'), 'rb') as f:
        shap_values = pickle.load(f)

    # Update kw_dict
    kw_dict = {
        'device': device,
        'num_items': num_items,
        'num_features': num_items,
        'demographic': False,
        'pop_array': np.array([pop_dict.get(i, 0) for i in range(num_items)]),
        'all_items_tensor': torch.eye(num_items).to(device),
        'static_test_data': test_data,
        'items_array': items_array,
        'output_type': output_type_dict[recommender_name],
        'recommender_name': recommender_name,
        'files_path': files_path
    }

    # Load recommender
    recommender = load_recommender(data_name, hidden_dim, checkpoints_path, recommender_path, **kw_dict)
    
    # Generate explanation dictionaries if they don't exist
    create_dictionaries = False  # Set to False if dictionaries already exist
    if create_dictionaries:
        recommender.eval()
        
        # Initialize dictionaries
        jaccard_expl_dict = {}
        cosine_expl_dict = {}
        lime_expl_dict = {}
        accent_expl_dict = {}
        shap_expl_dict = {}
        
        print(f"Generating explanation dictionaries for {data_name} {recommender_name}...")
        with torch.no_grad():
            for i in tqdm(range(test_array.shape[0])):
                user_vector = test_array[i]
                user_tensor = torch.FloatTensor(user_vector).to(device)
                user_id = int(test_data.index[i])

                item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
                item_vector = items_array[item_id]
                item_tensor = torch.FloatTensor(item_vector).to(device)

                user_vector[item_id] = 0
                user_tensor[item_id] = 0

                recommender.to(device)

                jaccard_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='jaccard')
                cosine_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='cosine')
                lime_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='lime')
                accent_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='accent')
                shap_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='shap', user_id=user_id)

        # Save dictionaries
        for name, dict_obj in [
            ('jaccard', jaccard_expl_dict),
            ('cosine', cosine_expl_dict),
            ('lime', lime_expl_dict),
            ('accent', accent_expl_dict),
            ('shap', shap_expl_dict)
        ]:
            with open(Path(files_path, f'{recommender_name}_{name}_expl_dict.pkl'), 'wb') as handle:
                pickle.dump(dict_obj, handle)
        
        print("Dictionaries generated and saved.")

    # Run all baselines
    baselines = ['jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']
    results = {}

    for baseline in baselines:
        print(f"Running {baseline} baseline for {data_name} {recommender_name}")
        results[baseline] = eval_one_expl_type(baseline)

    return results

def process_recommender(data_name, recommender_name):
    DP_DIR = Path("processed_data", data_name)
    files_path = Path("/storage/mikhail/PI4Rec", DP_DIR)
    
    num_users = num_users_dict[data_name]
    num_items = num_items_dict[data_name]
    num_features = num_items_dict[data_name]
    
    with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
        pop_dict = pickle.load(f)
    pop_array = np.zeros(len(pop_dict))
    for key, value in pop_dict.items():
        pop_array[key] = value

    test_data = pd.read_csv(Path(files_path,f'test_data_{data_name}.csv'), index_col=0)
    static_test_data = pd.read_csv(Path(files_path,f'static_test_data_{data_name}.csv'), index_col=0)
    
    test_array = static_test_data.iloc[:,:-2].to_numpy()
    items_array = np.eye(num_items)
    
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    all_items_tensor = torch.Tensor(items_array).to(device)

    output_type = output_type_dict[recommender_name]
    hidden_dim = hidden_dim_dict[(data_name,recommender_name)]
    recommender_path = recommender_path_dict[(data_name,recommender_name)]

    kw_dict = {
        'device': device,
        'num_items': num_items,
        'demographic': False,
        'num_features': num_features,
        'pop_array': pop_array,
        'all_items_tensor': all_items_tensor,
        'static_test_data': static_test_data,
        'items_array': items_array,
        'output_type': output_type,
        'recommender_name': recommender_name,
        'files_path': files_path
    }

    recommender = load_recommender(data_name, hidden_dim, checkpoints_path, recommender_path, **kw_dict)

    print(f"Processing {data_name} dataset with {recommender_name} recommender")
    
    results = {}
    for expl_name in ['pop', 'jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']:
        results[expl_name] = eval_one_expl_type(expl_name, data_name, recommender_name, test_array, test_data, items_array, recommender, kw_dict)
    
    if results:  # Check if results is not empty
        print(f"Got results for {data_name} {recommender_name}")
        print(f"Available metrics: {list(results.items())[0][1].keys()}")
        # plot_all_metrics(results, data_name, recommender_name) # Call commented out
    else:
        print(f"No results generated for {data_name} {recommender_name}")

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def save_results_to_excel(results, filename):
    wb = Workbook()
    
    # Create MF recommender sheet
    ws_mf = wb.active
    ws_mf.title = "MF Recommender"
    
    # Create VAE recommender sheet
    ws_vae = wb.create_sheet(title="VAE Recommender")
    
    for ws, title in [(ws_mf, "AUC values for explaining an MF recommender."), 
                      (ws_vae, "AUC values for explaining a VAE recommender.")]:
        
        # Add title
        ws['A1'] = f"Table: {title}"
        ws['A1'].font = Font(bold=True)
        ws.merge_cells('A1:G1')
        
        # Add headers
        headers = ['Method', 'k=5', 'k=10', 'k=20', 'DEL', 'INS', 'NDCG']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        # Add data
        for row, (method, values) in enumerate(results.items(), start=4):
            ws.cell(row=row, column=1, value=method)
            for col, value in enumerate(values, start=2):
                ws.cell(row=row, column=col, value=value)
    
    # Apply some styling
    for ws in [ws_mf, ws_vae]:
        for row in ws[f'A3:G{ws.max_row}']:
            for cell in row:
                cell.border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
    
    wb.save(filename)

def run_and_format_results(data_name, recommender_name):
    results = {}
    for expl_name in ['jaccard', 'cosine', 'lime', 'shap', 'accent', 'lxr']:
        raw_results = eval_one_expl_type(expl_name)
        
        # Extract POS values
        pos_at_5 = raw_results['POS_at_5'][-1]  # Last value represents 100% of items
        pos_at_10 = raw_results['POS_at_10'][-1]
        pos_at_20 = raw_results['POS_at_20'][-1]
        
        # Format results as per the desired output
        results[expl_name.upper()] = [
            pos_at_5,
            pos_at_10,
            pos_at_20,
            raw_results['DEL'][-1],
            raw_results['INS'][-1],
            raw_results['NDCG'][-1]
        ]
    
    return results

# Define datasets and recommenders
data_names = ["ML1M"]#, "Yahoo", "Pinterest"
recommender_names = ["MLP"]#, "VAE", "NCF"

# Create a mapping between explainer names and actual explainer functions
explainer_mapping = {
    'jaccard': find_jaccard_mask,
    'cosine': find_cosine_mask,
    'lime': explain_instance_lime,
    'lxr': find_lxr_mask,
    'accent': find_accent_mask,
    'shap': find_shapley_mask
}

# Store all results
all_results = {}

for data_name in data_names:
    # Setup paths and load data
    DP_DIR = Path("processed_data", data_name)
    files_path = Path("/storage/mikhail/PI4Rec", DP_DIR)
    
    # Get dataset dimensions
    num_users = num_users_dict[data_name] 
    num_items = num_items_dict[data_name] 
    num_features = num_items_dict[data_name]
        
    # Load popularity data
    with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
        pop_dict = pickle.load(f)
    pop_array = np.zeros(len(pop_dict))
    for key, value in pop_dict.items():
        pop_array[key] = value

    # Load training and test data
    train_data = pd.read_csv(Path(files_path, f'train_data_{data_name}.csv'), index_col=0)
    test_data = pd.read_csv(Path(files_path, f'test_data_{data_name}.csv'), index_col=0)
    static_test_data = pd.read_csv(Path(files_path, f'static_test_data_{data_name}.csv'), index_col=0)
    
    # Convert to arrays
    train_array = train_data.to_numpy()
    test_array = static_test_data.iloc[:,:-2].to_numpy()
    items_array = np.eye(num_items)
    all_items_tensor = torch.Tensor(items_array).to(device)

    for recommender_name in recommender_names:
        print(f"\n{'='*50}")
        print(f"Processing {data_name} dataset with {recommender_name} recommender")
        print(f"{'='*50}")
        
        # Setup recommender configuration
        output_type = output_type_dict[recommender_name]
        hidden_dim = hidden_dim_dict[(data_name, recommender_name)]
        recommender_path = recommender_path_dict[(data_name, recommender_name)]

        # Update kw_dict for current configuration
        kw_dict = {
            'device': device,
            'num_items': num_items,
            'demographic': False,
            'num_features': num_features,
            'pop_array': pop_array,
            'all_items_tensor': all_items_tensor,
            'static_test_data': static_test_data,
            'items_array': items_array,
            'output_type': output_type,
            'recommender_name': recommender_name,
            'files_path': files_path
        }

        try:
            # Run baselines and get results
            results = {}
            for baseline in ['jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']:
                print(f"Running {baseline} baseline for {data_name} {recommender_name}")
                results[baseline] = eval_one_expl_type(baseline)
            
            all_results[(data_name, recommender_name)] = results
            
            # Generate and save visualizations for current combination
            # plot_all_metrics(results, data_name, recommender_name) # Call commented out
            
        except Exception as e:
            print(f"Error processing {data_name}-{recommender_name}: {str(e)}")
            continue

# Create output directory
os.makedirs('plots', exist_ok=True)

print("\nAll evaluations completed successfully")

# plot_all_metrics(results, data_name, recommender_name) # Call commented out
