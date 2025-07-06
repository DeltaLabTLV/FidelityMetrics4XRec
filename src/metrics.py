import numpy as np
import torch
from tqdm import tqdm
from pathlib import Path
import pickle
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell

from src.utils import get_top_k, recommender_run, get_ndcg, get_user_recommended_item
from src.explainers import (find_jaccard_mask, find_cosine_mask, find_lime_mask, 
                            find_lire_mask, find_lxr_mask, find_accent_mask, 
                            find_shapley_mask, find_pop_mask)

def single_user_metrics(
    user_vector, user_tensor, item_id, item_tensor, recommender_model, expl_dict,
    metric_type='discrete', steps=5, mask_by='history', **kw_dict
):
    """
    Calculate metrics for a single user with gradual masking.
    """
    # Get sorted items by importance
    if isinstance(expl_dict, dict):
        POS_sim_items = list(sorted(expl_dict.items(), key=lambda item: item[1], reverse=True))
    else:
        POS_sim_items = expl_dict

    if metric_type == 'discrete':
        masking_steps = range(1, steps + 1)
    elif metric_type == 'continuous':
        masking_steps = np.linspace(0, 1, steps)
    else:
        raise ValueError(f"Unknown metric_type: {metric_type}")

    num_steps = len(masking_steps)
    
    # Initialize metric arrays
    POS_at_20 = [0] * num_steps
    DEL, INS, NDCG = [0] * num_steps, [0] * num_steps, [0] * num_steps

    for i, step in enumerate(masking_steps):
        if mask_by == 'history':
            if mask_by == 'history':
                base_len = int(user_tensor.sum().item())
            else: # explanation
                base_len = len(POS_sim_items)
        else: # explanation
            base_len = len(POS_sim_items)

        if metric_type == 'discrete':
            num_items_to_mask = step
        else: # continuous
            num_items_to_mask = int(np.round(step * base_len))
            if num_items_to_mask == 0 and step > 0:
                num_items_to_mask = 1

        # When masking is 100%, the user history should be empty
        # When masking is 100% (last step of continuous), the user history should be empty
        if metric_type == 'continuous' and i == num_steps - 1:
            masked_user_tensor = torch.zeros_like(user_tensor)
        else:
            # Create mask based on explanation
            mask = torch.zeros_like(user_tensor)
            for j in POS_sim_items[:num_items_to_mask]:
                idx = j[0] if isinstance(j, tuple) else j
                mask[idx] = 1
            masked_user_tensor = user_tensor * (1 - mask)

        # Get rankings
        POS_ranked_list = get_top_k(masked_user_tensor, user_tensor, recommender_model, **kw_dict)

        # Calculate rankings
        POS_index = list(POS_ranked_list.keys()).index(item_id) + 1 if item_id in POS_ranked_list else kw_dict['num_items']

        # Calculate metrics
        POS_at_20[i] = 1 if POS_index <= 20 else 0
        DEL[i] = float(recommender_run(masked_user_tensor, recommender_model, item_tensor, item_id, **kw_dict).detach().cpu().numpy())
        INS[i] = float(recommender_run(user_tensor - masked_user_tensor, recommender_model, item_tensor, item_id, **kw_dict).detach().cpu().numpy())
        NDCG[i] = get_ndcg(list(POS_ranked_list.keys()), item_id, **kw_dict)

    res = [DEL, INS, NDCG, POS_at_20]
    return [np.array(x) for x in res]


def eval_one_expl_type(
    expl_name, data_name, recommender_name, test_array, test_data, items_array, recommender, kw_dict, files_path,
    metric_type='discrete', steps=5, mask_by='history'
):
    """
    Evaluate one explanation type with gradual masking.
    """
    print(f' ============ Start explaining {data_name} {recommender_name} by {expl_name} ============')
    import os
    from pathlib import Path
    files_path = Path(files_path)

    # Load or generate explanations
    expl_dict_path = Path(files_path, f'{recommender_name}_explanation_dicts.pkl')

    if not expl_dict_path.exists():
        raise FileNotFoundError(f"Explanation dictionary not found at {expl_dict_path}")

    with open(expl_dict_path, 'rb') as handle:
        all_expl_dicts = pickle.load(handle)
    
    expl_dict = all_expl_dicts.get(expl_name)
    if expl_dict is None:
        raise ValueError(f"Explanation type '{expl_name}' not found in the dictionary.")

    # The logic for regenerating missing users is no longer needed with the new approach.
    # Explanations are now created in a single batch.

    recommender.eval()
    
    num_steps = steps
    users_DEL, users_INS, NDCG = np.zeros(num_steps), np.zeros(num_steps), np.zeros(num_steps)
    POS_at_20 = np.zeros(num_steps)

    with torch.no_grad():
        for i in tqdm(range(test_array.shape[0])):
            user_vector = test_array[i]
            user_tensor = torch.FloatTensor(user_vector).to(kw_dict['device'])
            user_id = int(test_data.index[i])
            item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
            item_tensor = torch.FloatTensor(items_array[item_id]).to(kw_dict['device'])
            user_vector[item_id] = 0
            user_tensor[item_id] = 0
            user_expl = expl_dict[user_id]

            res = single_user_metrics(
                user_vector, user_tensor, item_id, item_tensor, recommender, user_expl,
                metric_type=metric_type, steps=steps, mask_by=mask_by, **kw_dict
            )
            
            users_DEL += res[0]; users_INS += res[1]; NDCG += res[2]
            POS_at_20 += res[3]

    a = test_array.shape[0]
    results = {
        'DEL': users_DEL/a, 'INS': users_INS/a, 'NDCG': NDCG/a,
        'POS_at_20': POS_at_20/a
    }
    if metric_type == 'continuous':
        results['masking_percentages'] = np.linspace(0, 1, steps)
    return results

def create_results_table(results, data_name, recommender_name):
    """
    Create a comprehensive table of all metrics for each method and masking step
    """
    # Initialize the table structure
    table_data = []
    metrics = ['DEL', 'INS', 'NDCG', 'POS_at_20']
    
    for method in results.keys():
        # Determine the number of steps from the length of the first metric's data
        num_steps = len(next(iter(results[method].values())))
        
        for step in range(num_steps):
            row = {
                'Method': method.upper(),
                'Step': step + 1,  # 1-based indexing
                'Dataset': data_name,
                'Recommender': recommender_name
            }
            
            # Add all metrics for this method and step
            for metric in metrics:
                # Ensure the metric exists and has enough data for the current step
                if metric in results[method] and len(results[method][metric]) > step:
                    row[metric] = results[method][metric][step]
                else:
                    row[metric] = None  # Or some other placeholder
            
            table_data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(table_data)
    
    # Save to CSV
    csv_filename = f'results/tables/results_{data_name}_{recommender_name}.csv'
    df.to_csv(csv_filename, index=False)
    
    # Create and save a formatted Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = f"{data_name}_{recommender_name}_Results"
    
    # Add title (in row 1)
    ws['A1'] = f"Results for {data_name} dataset with {recommender_name} recommender"
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True)
    
    # Add headers (in row 3)
    headers = ['Method', 'Step', 'DEL', 'INS', 'NDCG', 'POS@20']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Add data with formatting
    current_method = None
    row_num = 4
    for _, row in df.iterrows():
        if current_method != row['Method']:
            current_method = row['Method']
            row_num += 1  # Add space between methods
        
        ws.cell(row=row_num, column=1, value=row['Method'])
        ws.cell(row=row_num, column=2, value=row['Step'])
        ws.cell(row=row_num, column=3, value=float(row['DEL']))
        ws.cell(row=row_num, column=4, value=float(row['INS']))
        ws.cell(row=row_num, column=5, value=float(row['NDCG']))
        ws.cell(row=row_num, column=6, value=float(row['POS_at_20']))
        
        row_num += 1
    
    # Apply formatting to all cells
    for row in ws.iter_rows(min_row=3, max_row=row_num-1):
        for cell in row:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if isinstance(cell.value, float):
                cell.number_format = '0.000'
    
    # Adjust column widths (skip merged cells)
    column_widths = {}
    for row in ws.iter_rows(min_row=3):  # Start from row 3 to skip merged cells
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            col = cell.column_letter
            width = len(str(cell.value)) + 2
            current_width = column_widths.get(col, 0)
            column_widths[col] = max(current_width, width)
    
    # Apply the calculated widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save Excel file
    excel_filename = f'results/tables/results_{data_name}_{recommender_name}.xlsx'
    wb.save(excel_filename)
    
    # Return DataFrame for further analysis if needed
    return df